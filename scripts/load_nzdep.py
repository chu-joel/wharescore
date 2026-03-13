"""
Load NZDep2023 meshblock data into PostGIS.
Reads the Excel file and inserts into the nzdep table.

Prerequisites:
    pip install openpyxl "psycopg[binary]"

Usage:
    python scripts/load_nzdep.py
"""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

import psycopg


# Configuration — change password to match your PostgreSQL install
DB_CONFIG = "dbname=wharescore user=postgres password=postgres host=localhost port=5432"

NZDEP_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data", "nzdep", "NZDep2023_MB2023.xlsx"
)


def load_nzdep():
    """Load NZDep2023 meshblock data from Excel into PostGIS."""

    print(f"Reading {NZDEP_FILE}...")
    wb = openpyxl.load_workbook(NZDEP_FILE, read_only=True)
    ws = wb["NZDep2023_MB2023"] if "NZDep2023_MB2023" in wb.sheetnames else wb.active

    # Read header row to find column positions
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Columns found: {headers}")

    # Find column indexes (case-insensitive matching)
    header_lower = [h.lower() if h else "" for h in headers]
    col_map = {}
    for i, h in enumerate(header_lower):
        if "mb2023" in h and "code" in h:
            col_map["mb2023_code"] = i
        elif h == "nzdep2023" or (h.startswith("nzdep2023") and "score" not in h):
            col_map["nzdep2023"] = i
        elif "score" in h:
            col_map["nzdep2023_score"] = i
        elif "sa1" in h and "code" in h:
            col_map["sa12023_code"] = i

    print(f"Column mapping: {col_map}")

    # Read all data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        mb_code = row[col_map["mb2023_code"]]
        if mb_code is None:
            continue

        # Ensure meshblock code is 7 characters with leading zeros
        mb_code = str(mb_code).zfill(7)

        nzdep = row[col_map.get("nzdep2023")]
        score = row[col_map.get("nzdep2023_score")]
        sa1_code = row[col_map.get("sa12023_code")]

        if sa1_code:
            sa1_code = str(sa1_code).zfill(9)

        rows.append((mb_code, nzdep, score, sa1_code))

    wb.close()
    print(f"Read {len(rows)} meshblock records")

    # Insert into database
    print(f"Connecting to database...")
    conn = psycopg.connect(DB_CONFIG)
    cur = conn.cursor()

    # Clear existing data
    cur.execute("DELETE FROM nzdep;")

    # Batch insert
    insert_sql = """
        INSERT INTO nzdep (mb2023_code, nzdep2023, nzdep2023_score, sa12023_code)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (mb2023_code) DO UPDATE SET
            nzdep2023 = EXCLUDED.nzdep2023,
            nzdep2023_score = EXCLUDED.nzdep2023_score,
            sa12023_code = EXCLUDED.sa12023_code;
    """

    batch_size = 1000
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        cur.executemany(insert_sql, batch)
        print(f"  Inserted {min(i + batch_size, len(rows))}/{len(rows)}...")

    conn.commit()

    # Verify
    cur.execute("SELECT COUNT(*) FROM nzdep;")
    count = cur.fetchone()[0]
    print(f"\nDone! {count} records loaded into nzdep table.")

    # Show sample
    cur.execute("""
        SELECT mb2023_code, nzdep2023, nzdep2023_score
        FROM nzdep
        WHERE nzdep2023 IS NOT NULL
        LIMIT 5;
    """)
    print("\nSample records:")
    for row in cur.fetchall():
        print(f"  MB {row[0]}: Deprivation {row[1]}, Score {row[2]}")

    cur.close()
    conn.close()


if __name__ == "__main__":
    load_nzdep()
