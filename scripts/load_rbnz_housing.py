"""
Load RBNZ M10 Housing data (hm10.xlsx) into PostGIS.
National aggregate quarterly series: house sales, HPI, housing stock value, residential investment.
Source: Reserve Bank of New Zealand / CoreLogic. Published 2026-01-13.
"""

import sys
import time
import openpyxl
import psycopg

DB = "postgresql://postgres:postgres@localhost:5432/wharescore"
XLSX_PATH = "data/bonds/hm10.xlsx"


def load_rbnz_housing(conn):
    print("\n=== RBNZ M10 HOUSING ===", flush=True)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=False)
    ws = wb["Data"]

    # Row 1: headers (None, 'House sales', 'House price index (HPI)', ...)
    # Row 5: series IDs
    # Row 6+: data (date, val, val, val, val)

    conn.execute("DROP TABLE IF EXISTS rbnz_housing CASCADE;")
    conn.execute("""
        CREATE TABLE rbnz_housing (
            quarter_end DATE PRIMARY KEY,
            house_sales INTEGER,
            house_price_index NUMERIC,
            housing_stock_value_m NUMERIC,
            residential_investment_real_m NUMERIC
        );
    """)

    rows_inserted = 0
    for row_idx in range(6, ws.max_row + 1):
        quarter_end = ws.cell(row=row_idx, column=1).value
        if quarter_end is None:
            continue
        house_sales = ws.cell(row=row_idx, column=2).value
        hpi = ws.cell(row=row_idx, column=3).value
        stock_value = ws.cell(row=row_idx, column=4).value
        res_invest = ws.cell(row=row_idx, column=5).value

        conn.execute(
            """INSERT INTO rbnz_housing
               (quarter_end, house_sales, house_price_index, housing_stock_value_m, residential_investment_real_m)
               VALUES (%s, %s, %s, %s, %s)""",
            (quarter_end, house_sales, hpi, stock_value, res_invest),
        )
        rows_inserted += 1

    wb.close()
    conn.commit()

    # Verify
    count = conn.execute("SELECT count(*) FROM rbnz_housing").fetchone()[0]
    date_range = conn.execute(
        "SELECT min(quarter_end), max(quarter_end) FROM rbnz_housing"
    ).fetchone()
    print(f"  Loaded {count} quarterly records ({date_range[0]} to {date_range[1]})")

    # Sample latest
    latest = conn.execute(
        "SELECT * FROM rbnz_housing ORDER BY quarter_end DESC LIMIT 1"
    ).fetchone()
    print(f"  Latest: {latest[0]} — {latest[1]:,} sales, HPI {latest[2]}, stock ${latest[3]:,}M")

    return count


def main():
    start = time.time()
    with psycopg.connect(DB) as conn:
        count = load_rbnz_housing(conn)

    elapsed = time.time() - start
    print(f"\nDone in {elapsed:.1f}s — {count} records loaded into rbnz_housing")


if __name__ == "__main__":
    main()
