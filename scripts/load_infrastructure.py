"""
Load Te Waihanga National Infrastructure Pipeline into PostGIS.

Reads all CSV and Excel files from data/infrastructure/, merges them,
and keeps the most recent record per project (by PrimaryKey).

Source: https://tewaihanga.govt.nz/the-pipeline/
Licence: CC BY 4.0

Schema changed between old CSVs (2023 Q3-2024 Q1, 28 cols) and
newer Excel files (2024 Q4+, 31 cols). This script normalises both
into a unified table.
"""

import csv
import os
import openpyxl
import psycopg

DB_CONN = "host=localhost dbname=wharescore user=postgres password=postgres"
DATA_DIR = r"D:\Projects\Experiments\propertyiq-poc\data\infrastructure"

# Quarter ordering for dedup: higher = more recent
QUARTER_ORDER = {}
for y in range(2023, 2027):
    for q in range(1, 5):
        QUARTER_ORDER[f"{y}-Q{q}"] = y * 10 + q

CREATE_TABLE = """
DROP TABLE IF EXISTS infrastructure_projects CASCADE;
CREATE TABLE infrastructure_projects (
    id SERIAL PRIMARY KEY,
    primary_key TEXT NOT NULL UNIQUE,
    submission_quarter TEXT,
    organisation TEXT,
    delivery_organisation TEXT,
    project_name TEXT NOT NULL,
    description TEXT,
    project_status TEXT,
    funding_status TEXT,
    procurement_type TEXT,
    procurement_method TEXT,
    region TEXT,
    city TEXT,
    suburb TEXT,
    sector TEXT,
    infra_sector TEXT,
    infra_service TEXT,
    infra_asset TEXT,
    value_range TEXT,
    quarter_bc_start TEXT,
    quarter_bc_end TEXT,
    quarter_procurement_start TEXT,
    quarter_procurement_end TEXT,
    quarter_construction_start TEXT,
    quarter_construction_end TEXT,
    quarter_project_start TEXT,
    quarter_project_end TEXT,
    info_url TEXT,
    contact TEXT,
    geom GEOMETRY(Point, 4326)
);
"""

CREATE_INDEXES = """
CREATE INDEX idx_infra_geom ON infrastructure_projects USING GIST (geom);
CREATE INDEX idx_infra_region ON infrastructure_projects (region);
CREATE INDEX idx_infra_sector ON infrastructure_projects (sector);
CREATE INDEX idx_infra_status ON infrastructure_projects (project_status);
CREATE INDEX idx_infra_city ON infrastructure_projects (city);
CREATE INDEX idx_infra_suburb ON infrastructure_projects (suburb);
CREATE INDEX idx_infra_quarter ON infrastructure_projects (submission_quarter);
"""


def clean(val):
    """Clean a value: strip whitespace, convert None-like strings to None."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None", "none", "N/A", "n/a"):
        return None
    return s


def make_geom(lat_str, lng_str):
    """Create EWKT point from lat/lng strings, or None."""
    lat = clean(lat_str)
    lng = clean(lng_str)
    if lat and lng:
        try:
            return f"SRID=4326;POINT({float(lng)} {float(lat)})"
        except ValueError:
            pass
    return None


def parse_csv(filepath):
    """Parse old-format CSV (2023 Q3 - 2024 Q1)."""
    records = {}
    with open(filepath, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pk = clean(row["PrimaryKey"])
            if not pk:
                continue

            # Derive quarter from DateUpdated (YYYY-MM-DD)
            date_str = clean(row.get("DateUpdated", ""))
            quarter = None
            if date_str and len(date_str) >= 7:
                try:
                    y = int(date_str[:4])
                    m = int(date_str[5:7])
                    q = (m - 1) // 3 + 1
                    quarter = f"{y}-Q{q}"
                except ValueError:
                    pass

            rec = {
                "primary_key": pk,
                "submission_quarter": quarter,
                "organisation": clean(row.get("ProcuringAgencyOrganisation")),
                "delivery_organisation": None,
                "project_name": clean(row.get("ProjectName")) or "Unknown",
                "description": clean(row.get("ProjectShortDescription")),
                "project_status": clean(row.get("ProjectStatus")),
                "funding_status": clean(row.get("FundingStatus")),
                "procurement_type": clean(row.get("ProcurementType")),
                "procurement_method": clean(row.get("ProcurementMethod")),
                "region": clean(row.get("ProjectRegion")),
                "city": clean(row.get("ProjectCityTown")),
                "suburb": clean(row.get("ProjectSuburb")),
                "sector": clean(row.get("ProjectSector")),
                "infra_sector": None,
                "infra_service": None,
                "infra_asset": None,
                "value_range": clean(row.get("EstimatedProjectValueRange")),
                "quarter_bc_start": clean(row.get("EstimatedQuarterBusinessCaseStart")),
                "quarter_bc_end": clean(row.get("EstimatedQuarterBusinessCaseCompletion")),
                "quarter_procurement_start": clean(row.get("EstimatedQuarterProcurementStart")),
                "quarter_procurement_end": clean(row.get("EstimatedQuarterProcurementCompletion")),
                "quarter_construction_start": clean(row.get("EstimatedQuarterConstructionStart")),
                "quarter_construction_end": clean(row.get("EstimatedQuarterConstructionCompletion")),
                "quarter_project_start": clean(row.get("EstimatedQuarterProjectRangeStart")),
                "quarter_project_end": clean(row.get("EstimatedQuarterProjectRangeCompletion")),
                "info_url": clean(row.get("ProjectInfoURL")),
                "contact": clean(row.get("Contact")),
                "geom": make_geom(row.get("Latitude"), row.get("Longitude")),
            }

            # Keep most recent version per primary key
            existing = records.get(pk)
            if existing is None:
                records[pk] = rec
            else:
                eq = existing.get("submission_quarter") or ""
                rq = rec.get("submission_quarter") or ""
                if QUARTER_ORDER.get(rq, 0) > QUARTER_ORDER.get(eq, 0):
                    records[pk] = rec

    return records


def parse_xlsx(filepath):
    """Parse new-format Excel (2024 Q4+)."""
    records = {}
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb["pipeline-data"] if "pipeline-data" in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col_idx = {name: i for i, name in enumerate(header)}

    def get(row, col_name):
        idx = col_idx.get(col_name)
        if idx is None:
            return None
        return row[idx]

    for row in rows_iter:
        pk = clean(get(row, "PrimaryKey"))
        if not pk:
            continue

        rec = {
            "primary_key": pk,
            "submission_quarter": clean(get(row, "InformationSubmissionQuarter")),
            "organisation": clean(get(row, "InformationSubmissionOrganisation")),
            "delivery_organisation": clean(get(row, "DeliveryOrganisation")),
            "project_name": clean(get(row, "ProjectName")) or "Unknown",
            "description": clean(get(row, "ProjectDescription")),
            "project_status": clean(get(row, "ProjectStatus")),
            "funding_status": clean(get(row, "FundingStatus")),
            "procurement_type": clean(get(row, "ProcurementType")),
            "procurement_method": clean(get(row, "ProcurementMethod")),
            "region": clean(get(row, "ProjectRegion")),
            "city": clean(get(row, "ProjectCityTown")),
            "suburb": clean(get(row, "ProjectSuburbOrPostcode")),
            "sector": clean(get(row, "SubmittedProjectSector")),
            "infra_sector": clean(get(row, "CommissionAssignedInfrastructureSector")),
            "infra_service": clean(get(row, "CommissionAssignedInfrastructureService")),
            "infra_asset": clean(get(row, "CommissionAssignedInfrastructureAsset")),
            "value_range": clean(get(row, "EstimatedProjectValueRange")),
            "quarter_bc_start": clean(get(row, "EstimatedQuarterBusinessCaseStart")),
            "quarter_bc_end": clean(get(row, "EstimatedQuarterBusinessCaseCompletion")),
            "quarter_procurement_start": clean(get(row, "EstimatedQuarterProcurementStart")),
            "quarter_procurement_end": clean(get(row, "EstimatedQuarterProcurementCompletion")),
            "quarter_construction_start": clean(get(row, "EstimatedQuarterConstructionStart")),
            "quarter_construction_end": clean(get(row, "EstimatedQuarterConstructionCompletion")),
            "quarter_project_start": clean(get(row, "EstimatedQuarterProjectRangeStart")),
            "quarter_project_end": clean(get(row, "EstimatedQuarterProjectRangeCompletion")),
            "info_url": clean(get(row, "ProjectInfoURL")),
            "contact": clean(get(row, "Contact")),
            "geom": make_geom(get(row, "Latitude"), get(row, "Longitude")),
        }

        existing = records.get(pk)
        if existing is None:
            records[pk] = rec
        else:
            eq = existing.get("submission_quarter") or ""
            rq = rec.get("submission_quarter") or ""
            if QUARTER_ORDER.get(rq, 0) > QUARTER_ORDER.get(eq, 0):
                records[pk] = rec

    wb.close()
    return records


def load():
    # Collect all records, deduplicating by primary key (most recent wins)
    all_records = {}
    files_loaded = []

    for fname in sorted(os.listdir(DATA_DIR)):
        fpath = os.path.join(DATA_DIR, fname)
        if fname.startswith("~"):
            continue

        if fname.endswith(".csv"):
            print(f"Reading CSV: {fname}")
            records = parse_csv(fpath)
        elif fname.endswith(".xlsx"):
            print(f"Reading Excel: {fname}")
            records = parse_xlsx(fpath)
        else:
            continue

        files_loaded.append(f"{fname} ({len(records)} records)")

        # Merge: keep most recent per primary key
        for pk, rec in records.items():
            existing = all_records.get(pk)
            if existing is None:
                all_records[pk] = rec
            else:
                eq = existing.get("submission_quarter") or ""
                rq = rec.get("submission_quarter") or ""
                if QUARTER_ORDER.get(rq, 0) > QUARTER_ORDER.get(eq, 0):
                    all_records[pk] = rec

    print(f"\nFiles loaded: {len(files_loaded)}")
    for f in files_loaded:
        print(f"  {f}")
    print(f"Unique projects after dedup: {len(all_records)}")

    # Load into PostGIS
    with psycopg.connect(DB_CONN) as conn:
        with conn.cursor() as cur:
            print("\nCreating table...")
            cur.execute(CREATE_TABLE)

            print("Inserting records...")
            insert_sql = """
                INSERT INTO infrastructure_projects (
                    primary_key, submission_quarter, organisation, delivery_organisation,
                    project_name, description, project_status, funding_status,
                    procurement_type, procurement_method, region, city, suburb,
                    sector, infra_sector, infra_service, infra_asset, value_range,
                    quarter_bc_start, quarter_bc_end, quarter_procurement_start,
                    quarter_procurement_end, quarter_construction_start,
                    quarter_construction_end, quarter_project_start, quarter_project_end,
                    info_url, contact, geom
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
            """

            rows = []
            for rec in all_records.values():
                rows.append((
                    rec["primary_key"], rec["submission_quarter"],
                    rec["organisation"], rec["delivery_organisation"],
                    rec["project_name"], rec["description"],
                    rec["project_status"], rec["funding_status"],
                    rec["procurement_type"], rec["procurement_method"],
                    rec["region"], rec["city"], rec["suburb"],
                    rec["sector"], rec["infra_sector"],
                    rec["infra_service"], rec["infra_asset"],
                    rec["value_range"],
                    rec["quarter_bc_start"], rec["quarter_bc_end"],
                    rec["quarter_procurement_start"], rec["quarter_procurement_end"],
                    rec["quarter_construction_start"], rec["quarter_construction_end"],
                    rec["quarter_project_start"], rec["quarter_project_end"],
                    rec["info_url"], rec["contact"], rec["geom"],
                ))

            cur.executemany(insert_sql, rows)
            print(f"Inserted {len(rows)} projects")

            print("Creating indexes...")
            cur.execute(CREATE_INDEXES)
            conn.commit()

            # Validation
            cur.execute("SELECT COUNT(*) FROM infrastructure_projects")
            total = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM infrastructure_projects WHERE geom IS NOT NULL")
            with_geom = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM infrastructure_projects WHERE region ILIKE '%Wellington%'")
            wellington = cur.fetchone()[0]
            cur.execute("""
                SELECT COUNT(DISTINCT submission_quarter) FROM infrastructure_projects
            """)
            quarters = cur.fetchone()[0]
            cur.execute("""
                SELECT project_name, project_status, value_range, sector,
                       city, suburb, submission_quarter,
                       CASE WHEN geom IS NOT NULL THEN 'yes' ELSE 'no' END as has_coords
                FROM infrastructure_projects
                WHERE region ILIKE '%Wellington%'
                ORDER BY value_range DESC
                LIMIT 15
            """)
            wgtn_projects = cur.fetchall()

            print(f"\n--- Validation ---")
            print(f"Total unique projects: {total}")
            print(f"With geometry: {with_geom}")
            print(f"Wellington region: {wellington}")
            print(f"Quarters covered: {quarters}")
            print(f"\nTop Wellington projects by value:")
            for p in wgtn_projects:
                print(f"  {p[0]} | {p[1]} | {p[2]} | {p[3]} | {p[4] or ''} {p[5] or ''} | Q:{p[6]} | coords:{p[7]}")

    # Analyze table
    with psycopg.connect(DB_CONN) as conn:
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute("ANALYZE infrastructure_projects")
    print("\nTable analyzed. Done!")


if __name__ == "__main__":
    load()
