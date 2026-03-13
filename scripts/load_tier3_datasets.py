"""Load Tier 3 datasets into PostGIS: air quality sites, water quality sites, heritage list, wildfire risk."""
import csv
import json
import subprocess
import sys
import os

PSQL = "E:/Programs/postgresql/bin/psql.exe"
DB = "wharescore"
os.environ["PGPASSWORD"] = "postgres"


def run_sql(sql, quiet=False):
    cmd = [PSQL, "-U", "postgres", "-d", DB, "-c", sql]
    if quiet:
        cmd.insert(-2, "-q")
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0 and not quiet:
        print(f"  SQL ERROR: {result.stderr.strip()}", flush=True)
    return result


def load_sql_file(path):
    cmd = [PSQL, "-U", "postgres", "-d", DB, "-q", "-f", path]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"  LOAD ERROR: {result.stderr.strip()}", flush=True)
    return result


# ============================================================
# 1. AIR QUALITY MONITORING SITES
# ============================================================
def load_air_quality():
    """Extract unique air quality sites from LAWA Excel and load as point table."""
    print("\n=== Loading Air Quality Sites ===", flush=True)
    import openpyxl

    wb = openpyxl.load_workbook(
        "D:/Projects/Experiments/propertyiq-poc/data/air-quality/lawa-air-quality-2016-2024.xlsx",
        read_only=True,
    )
    ws = wb["Monitoring dataset "]

    # Extract unique sites
    sites = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i == 0:
            continue  # skip if header repeated
        region, agency, town, site_name, lawa_id, site_id, lat, lon, site_type, indicator, sample_date, conc = row
        if lawa_id and lat and lon and lawa_id not in sites:
            sites[lawa_id] = {
                "region": region,
                "agency": agency,
                "town": town,
                "site_name": site_name,
                "lawa_id": lawa_id,
                "site_type": site_type,
                "lat": float(lat),
                "lon": float(lon),
            }
    wb.close()

    # Also get trend data per site
    wb = openpyxl.load_workbook(
        "D:/Projects/Experiments/propertyiq-poc/data/air-quality/lawa-air-quality-2016-2024.xlsx",
        read_only=True,
    )
    ws = wb["Ten-year trends"]
    trends = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        region, agency, town, site_name, lawa_id, indicator, year, trend_result = row
        if lawa_id and year == 2024:
            key = (lawa_id, indicator)
            trends[key] = trend_result
    wb.close()

    print(f"  Found {len(sites)} unique air quality monitoring sites", flush=True)

    # Create table and insert
    run_sql("DROP TABLE IF EXISTS air_quality_sites;")
    run_sql("""
        CREATE TABLE air_quality_sites (
            id SERIAL PRIMARY KEY,
            lawa_id TEXT NOT NULL UNIQUE,
            site_name TEXT,
            town TEXT,
            region TEXT,
            agency TEXT,
            site_type TEXT,
            pm10_trend TEXT,
            pm25_trend TEXT,
            geom GEOMETRY(Point, 4326)
        );
    """)

    for s in sites.values():
        pm10_trend = trends.get((s["lawa_id"], "PM10"), None)
        pm25_trend = trends.get((s["lawa_id"], "PM2.5"), None)
        pm10_str = f"'{pm10_trend}'" if pm10_trend else "NULL"
        pm25_str = f"'{pm25_trend}'" if pm25_trend else "NULL"
        name = s["site_name"].replace("'", "''") if s["site_name"] else ""
        town = s["town"].replace("'", "''") if s["town"] else ""
        region = s["region"].replace("'", "''") if s["region"] else ""
        agency = s["agency"].replace("'", "''") if s["agency"] else ""
        stype = s["site_type"].replace("'", "''") if s["site_type"] else ""
        run_sql(
            f"""INSERT INTO air_quality_sites (lawa_id, site_name, town, region, agency, site_type, pm10_trend, pm25_trend, geom)
            VALUES ('{s["lawa_id"]}', '{name}', '{town}', '{region}', '{agency}', '{stype}',
                    {pm10_str}, {pm25_str},
                    ST_SetSRID(ST_MakePoint({s["lon"]}, {s["lat"]}), 4326));""",
            quiet=True,
        )

    run_sql("CREATE INDEX idx_air_quality_geom ON air_quality_sites USING GIST(geom);")
    run_sql("ANALYZE air_quality_sites;")

    result = run_sql("SELECT COUNT(*) FROM air_quality_sites;")
    print(f"  Loaded: {result.stdout.strip()}", flush=True)


# ============================================================
# 2. WATER QUALITY MONITORING SITES
# ============================================================
def load_water_quality():
    """Extract unique river water quality sites with state bands from LAWA Excel."""
    print("\n=== Loading Water Quality Sites ===", flush=True)
    import openpyxl

    wb = openpyxl.load_workbook(
        "D:/Projects/Experiments/propertyiq-poc/data/water-quality/lawa-river-state-trend-2025.xlsx",
        read_only=True,
    )

    # Get state attribute bands (A-E grades per indicator)
    ws = wb["State Attribute Band"]
    sites = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        (region, agency, catchment, lawa_id, site_id, council_id, hyear,
         lat, lon, sed_class, landcover, altitude, rec_landcover,
         indicator, units, band, median, p95, rec260, rec540) = row
        if lawa_id and lat and lon and str(lat) != "NA" and str(lon) != "NA":
            try:
                lat_f, lon_f = float(lat), float(lon)
            except (ValueError, TypeError):
                continue
            if lawa_id not in sites:
                sites[lawa_id] = {
                    "region": region,
                    "agency": agency,
                    "catchment": catchment,
                    "lawa_id": lawa_id,
                    "site_name": site_id,
                    "lat": lat_f,
                    "lon": lon_f,
                    "bands": {},
                }
            # Store the worst (latest year) band per indicator
            if indicator:
                ind_key = indicator.split(" / ")[0].strip() if " / " in indicator else indicator
                sites[lawa_id]["bands"][ind_key] = band
    wb.close()

    print(f"  Found {len(sites)} unique water quality monitoring sites", flush=True)

    run_sql("DROP TABLE IF EXISTS water_quality_sites;")
    run_sql("""
        CREATE TABLE water_quality_sites (
            id SERIAL PRIMARY KEY,
            lawa_id TEXT NOT NULL UNIQUE,
            site_name TEXT,
            catchment TEXT,
            region TEXT,
            agency TEXT,
            ecoli_band TEXT,
            ammonia_band TEXT,
            nitrate_band TEXT,
            drp_band TEXT,
            clarity_band TEXT,
            geom GEOMETRY(Point, 4326)
        );
    """)

    for s in sites.values():
        def band_val(key):
            v = s["bands"].get(key)
            return f"'{v}'" if v else "NULL"

        name = str(s["site_name"] or "").replace("'", "''")
        catchment = str(s["catchment"] or "").replace("'", "''")
        region = str(s["region"] or "").replace("'", "''")
        agency = str(s["agency"] or "").replace("'", "''")
        run_sql(
            f"""INSERT INTO water_quality_sites (lawa_id, site_name, catchment, region, agency,
                ecoli_band, ammonia_band, nitrate_band, drp_band, clarity_band, geom)
            VALUES ('{s["lawa_id"]}', '{name}', '{catchment}', '{region}', '{agency}',
                    {band_val("E. coli")}, {band_val("Ammonical nitrogen")},
                    {band_val("Nitrate")}, {band_val("Dissolved reactive phosphorus")},
                    {band_val("Clarity")},
                    ST_SetSRID(ST_MakePoint({s["lon"]}, {s["lat"]}), 4326));""",
            quiet=True,
        )

    run_sql("CREATE INDEX idx_water_quality_geom ON water_quality_sites USING GIST(geom);")
    run_sql("ANALYZE water_quality_sites;")

    result = run_sql("SELECT COUNT(*) FROM water_quality_sites;")
    print(f"  Loaded: {result.stdout.strip()}", flush=True)


# ============================================================
# 3. HERITAGE NZ LIST
# ============================================================
def load_heritage():
    """Load Heritage NZ list from Algolia JSON (has coordinates)."""
    print("\n=== Loading Heritage NZ List ===", flush=True)

    with open("D:/Projects/Experiments/propertyiq-poc/data/heritage/heritage-nz-algolia.json") as f:
        records = json.load(f)

    print(f"  {len(records)} records with coordinates", flush=True)

    run_sql("DROP TABLE IF EXISTS heritage_sites;")
    run_sql("""
        CREATE TABLE heritage_sites (
            id SERIAL PRIMARY KEY,
            list_number INTEGER NOT NULL UNIQUE,
            name TEXT NOT NULL,
            list_entry_type TEXT,
            list_entry_status TEXT,
            address TEXT,
            district_council TEXT,
            region TEXT,
            geom GEOMETRY(Point, 4326)
        );
    """)

    # Write SQL file and load with psql -f (avoids Windows command line length limits)
    sql_path = "D:/Projects/Experiments/propertyiq-poc/data/heritage/heritage_load.sql"
    with open(sql_path, "w", encoding="utf-8") as f:
        for r in records:
            loc = r.get("location", {})
            lat = loc.get("latitude")
            lon = loc.get("longitude")
            if not lat or not lon:
                continue
            name = r.get("name", "").replace("'", "''")
            address = (r.get("address") or "").replace("'", "''")
            entry_type = (r.get("listEntryType") or "").replace("'", "''")
            status = (r.get("listEntryStatus") or "").replace("'", "''")
            district = (r.get("districtCouncil") or "").replace("'", "''")
            region = (r.get("region") or "").replace("'", "''")
            f.write(
                f"INSERT INTO heritage_sites (list_number, name, list_entry_type, list_entry_status, "
                f"address, district_council, region, geom) VALUES "
                f"({r['listNumber']}, '{name}', '{entry_type}', '{status}', "
                f"'{address}', '{district}', '{region}', "
                f"ST_SetSRID(ST_MakePoint({lon}, {lat}), 4326));\n"
            )
    load_sql_file(sql_path)

    run_sql("CREATE INDEX idx_heritage_geom ON heritage_sites USING GIST(geom);")
    run_sql("CREATE INDEX idx_heritage_type ON heritage_sites(list_entry_type);")
    run_sql("ANALYZE heritage_sites;")

    result = run_sql("SELECT COUNT(*) FROM heritage_sites;")
    print(f"  Loaded: {result.stdout.strip()}", flush=True)


# ============================================================
# 4. WILDFIRE RISK STATIONS
# ============================================================
def load_wildfire():
    """Load wildfire risk stations from Stats NZ CSVs."""
    print("\n=== Loading Wildfire Risk Stations ===", flush=True)

    run_sql("DROP TABLE IF EXISTS wildfire_risk;")
    run_sql("""
        CREATE TABLE wildfire_risk (
            id SERIAL PRIMARY KEY,
            site TEXT NOT NULL,
            fuel_type TEXT NOT NULL,
            ten_year_mean REAL,
            quantile TEXT,
            slope_decade REAL,
            trend_likelihood TEXT,
            geom GEOMETRY(Point, 4326),
            UNIQUE(site, fuel_type)
        );
    """)

    for fuel_file in ["forest-vhe-wildfire-risk.csv", "grass-vhe-wildfire-risk.csv"]:
        path = f"D:/Projects/Experiments/propertyiq-poc/data/wildfire/{fuel_file}"
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                site = row["site"].replace("'", "''").strip()
                fuel = row["fuel_type"].strip()
                mean = row["ten_year_mean"] if row["ten_year_mean"] != "NA" else "NULL"
                quantile = row.get("quantile", "").replace("\u2013", "-")
                slope = row.get("slope_decade", "NA")
                slope = slope if slope != "NA" else "NULL"
                trend = row.get("trend_likelihood", "")
                trend_str = f"'{trend}'" if trend and trend != "NA" else "NULL"
                lat = row["lat"]
                lon = row["lon"]
                run_sql(
                    f"""INSERT INTO wildfire_risk (site, fuel_type, ten_year_mean, quantile, slope_decade, trend_likelihood, geom)
                    VALUES ('{site}', '{fuel}', {mean}, '{quantile}', {slope}, {trend_str},
                            ST_SetSRID(ST_MakePoint({lon}, {lat}), 4326))
                    ON CONFLICT (site, fuel_type) DO NOTHING;""",
                    quiet=True,
                )

    run_sql("CREATE INDEX idx_wildfire_geom ON wildfire_risk USING GIST(geom);")
    run_sql("ANALYZE wildfire_risk;")

    result = run_sql("SELECT COUNT(*) FROM wildfire_risk;")
    print(f"  Loaded: {result.stdout.strip()}", flush=True)


if __name__ == "__main__":
    load_air_quality()
    load_water_quality()
    load_heritage()
    load_wildfire()
    print("\n=== All Tier 3 datasets loaded! ===", flush=True)
